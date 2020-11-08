import openpyxl
from openpyxl.styles import Alignment
import pandas as pd
file=["A&NIslands", "AndhraPradeshOld", "ArunachalPradesh", "Assam", "Bihar", "Chandigarh", "Chhattisgarh"]
year="2010-11"
for j in range(7):
	filename=file[j]+year+".xlsx"
	book= openpyxl.load_workbook(filename)
	sheet=book["Sheet1"]
	sheet.unmerge_cells('A7:B9')
	sheet.delete_rows(9)
	sheet.delete_rows(1,7)
	i=326
	columns={2,3,5,13,15,21,25,27,33,35,37,45,47,51,55,63,65,89,93,101,105,113,125,127,143,145,147,187,189,201,203,207,209,215,217,219,227,231,241,277,279,309}
	while(i>0):
		if i in columns:
	 	 i=i-1
	 	 continue
		else:
	 	 sheet.delete_cols(i)
	 	 i=i-1
	sheet.cell(row=1,column=2).alignment=Alignment(horizontal='left')
	sheet.cell(row=1,column=1).value="Indicator"
	book.save(filename)
	read_file=pd.read_excel(filename)
	filename_csv=file[j]+year+".csv"
	read_file.to_csv(filename_csv,index=None,header=True)
